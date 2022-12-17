from django.db.utils import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.status import (
        HTTP_201_CREATED as ST_201,
        HTTP_204_NO_CONTENT as ST_204,
        HTTP_400_BAD_REQUEST as ST_400,
        HTTP_401_UNAUTHORIZED as ST_401,
        HTTP_409_CONFLICT as ST_409
)

from base.perms import UserIsStaff
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from .resources import CensusResource
from tablib import Dataset
from .models import Census
from django.views.generic import ListView
from django.http import HttpResponse
from django.shortcuts import render
from .forms import CreationCensusForm
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def filter(request):
    censo = Census.objects.all()
    votingsIds = votingIdSet()
    return render(request, 'filterCensus.html', 
        {'census' : censo, 'votingsIds': votingsIds})

class FilterVotingID(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('i')
        return Census.objects.filter(voting_id__icontains=query).order_by('-voting_id')

class FilterVoterID(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('i')
        return Census.objects.filter(voting_id__icontains=query).order_by('-voter_id')

class FilterName(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('q')
        return Census.objects.filter(name__icontains=query).order_by('-name')

class FilterSurname(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('i')
        return Census.objects.filter(surname__icontains=query).order_by('-surname')

class FilterCity(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('j')
        return Census.objects.filter(city__icontains=query).order_by('-city')

class FilterCommunity(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('j')
        return Census.objects.filter(a_community__icontains=query).order_by('-a_community')

class FilterGender(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('j')
        return Census.objects.filter(gender__icontains=query).order_by('-gender')

class FilterBornYear(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('j')
        return Census.objects.filter(born_year__icontains=query).order_by('-born_year')

class FilterCivilState(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('j')
        return Census.objects.filter(civil_state__icontains=query).order_by('-civil_state')

class FilterSexuality(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('j')
        return Census.objects.filter(sexuality__icontains=query).order_by('-sexuality')

class FilterWorks(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'

    def get_queryset(self):
        query = self.request.GET.get('j')
        return Census.objects.filter(works__icontains=query).order_by('-works')
    

def importer(request):
    try:
        if request.method == 'POST':
            census_resource = CensusResource()
            # obtendremos datos en census_resource
            dataset = Dataset()
            new_census = request.FILES['myfile']

            if not new_census.name.endswith('xlsx'):
                messages.error(request, 'incorrect format, must be .xlsx')
                return render(request, 'importer.html')
            messages.info(request, 'Uploading Data Line by Line...')

            imported_data = dataset.load(new_census.read(),format='xlsx')
            #print(imported_data)
            count = 1
            for data in imported_data:
                value = Census(
                        data[0],
                        data[1],
                        data[2],
                        data[3],
                        data[4],
                        data[5],
                        data[6],
                        data[7],
                        data[8],
                        data[9],
                        data[10],
                        data[11]
                        )
                value.save()
            # time.sleep(1)
            messages.info(request, 'File Uploaded Successfully...')
            
            #result = census_resource.import_data(dataset, dry_run=True)  # Test the data import

            #if not result.has_errors():
            #    census_resource.import_data(dataset, dry_run=False)  # Actually import now
     
    except:
        messages.error(request,'Same voter_id has been observed more than once. Import has been canceled../nPlease Make sure voter_id field should be unique.')
    
    return render(request, 'importer.html')


class CensusCreate(generics.ListCreateAPIView):
    permission_classes = (UserIsStaff,)

    def create(self, request, *args, **kwargs):
        voting_id = request.data.get('voting_id')
        voters = request.data.get('voters')
        try:
            for voter in voters:
                census = Census(voting_id=voting_id, voter_id=voter)
                census.save()
        except IntegrityError:
            return Response('Error try to create census', status=ST_409)
        return Response('Census created', status=ST_201)

    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get('voting_id')
        voters = Census.objects.filter(voting_id=voting_id).values_list('voter_id', flat=True)
        return Response({'voters': voters})


class CensusDetail(generics.RetrieveDestroyAPIView):


    def destroy(self, request, voting_id, *args, **kwargs):
        voters = request.data.get('voters')
        census = Census.objects.filter(voting_id=voting_id, voter_id__in=voters)
        census.delete()
        return Response('Voters deleted from census', status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        voter = request.GET.get('voter_id')
        try:
            Census.objects.get(voting_id=voting_id)
        except ObjectDoesNotExist:
            return Response('Invalid voter', status=ST_401)
        return Response('Valid voter')




def GetId(request):
    id = request.GET['id']
    
    census = Census.objects.filter(voting_id=int(id))
    if len(census) == 0:
        return render(request,'census.html',{'error_id':'There is not a census with that voting_id'})
    else:
        return render(request,"census_details.html",{'census':census})


def hello(request):
    return render(request,'census.html')

######Creación de censo
def createCensus(request): 
    if request.method == 'GET':
        return render(request, 'census_create.html',{'form': CreationCensusForm})
    else: 
        if request.method == 'POST':
            try: 
                census = Census.objects.create(voting_id = request.POST['voting_id'],voter_id = request.POST['voter_id'],
                name = request.POST['name'],surname= request.POST['surname'],city = request.POST['city'],a_community = request.POST['a_community'],
                gender = request.POST['gender'],born_year = request.POST['born_year'],civil_state = request.POST['civil_state'],
                sexuality = request.POST['sexuality'],works = request.POST['works'])
                census.save()
                return render(request,'census_succeed.html',{'census':census})
                
            except: 
                return render(request,'census_create.html',{'form': CreationCensusForm, "error": 'Census already exist'})
        return  render(request,'census_create.html',{'form': CreationCensusForm})


############BORRAR CENSOS
def deleteCensus(request):
    Voterid = request.GET['Voterid']
    Votingid = request.GET['Votingid']
    census = Census.objects.filter(voting_id=int(Votingid),voter_id = int(Voterid))
    if len(census) == 0: 
        return render(request,'census.html',{'error':'Census does not exist.Try other census'})
    if len(census)==1:
        census.delete()
        return render(request,'census_deleted.html')
    

def censusCreatedSucced(request):
    return render(request,'census_succeed.html')


def censusDeletedSucced(request):
    return render(request,'census_deleted.html')

def home(request):
    queryset = Census.objects.all()
    return render(request, 'lista_censo.html', {'queryset':queryset})


'''
class ReportePersonalizadoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        query = Census.objects.all()
        wb = Workbook()
        bandera = True
        cont = 1
        for q in query:
            if bandera:
                ws = wb.active
                bandera = False
            else:
                ws = wb.create_sheet('Hoja'+str(cont))
            cont += 1

        #Establecer el nombre de mi archivo
        nombre_archivo = "ReportePersonalizadoExcel.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
'''

#Exportar censo en excel

class ReporteAutorExcel(TemplateView):

    def get(self,request,*args,**kwargs):

        census = Census.objects.all()
        wb = Workbook()
        ws = wb.active
        

        ws.merge_cells('B1:L1')

        ws['B3'] = 'Voting_id'
        ws['C3'] = 'Voter_id'
        ws['D3'] = 'Name'
        ws['E3'] = 'Surname'
        ws['F3'] = 'City'
        ws['G3'] = 'A_community'
        ws['H3'] = 'Gender'
        ws['I3'] = 'Born_year'
        ws['J3'] = 'Civil_state'
        ws['K3'] = 'Sexuality'
        ws['L3'] = 'Works'

        cont = 4

        for censu in census:
            ws.cell(row = cont, column = 2).value = censu.voting_id
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 2).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 3).value = censu.voter_id
            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 3).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 4).value = censu.name
            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 4).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 5).value = censu.surname
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 5).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 6).value = censu.city
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 6).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 7).value = censu.a_community
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 7).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 8).value = censu.gender
            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 8).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 9).value = censu.born_year
            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 9).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 10).value = censu.civil_state
            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 10).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 11).value = censu.sexuality
            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 11).font = Font(name = 'Calibri', size = 8)

            ws.cell(row = cont, column = 12).value = censu.works
            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 12).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))
            ws.cell(row = cont, column = 12).font = Font(name = 'Calibri', size = 8)

            ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))                         
            ws['B1'].fill = PatternFill(start_color = '66FFCC', end_color ='66FFCC', fill_type = "solid") 
            ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
            ws['B1'] = 'EXPORTACIÓN DE CENSOS'

            ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))  
            ws['J3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['J3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['K3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['K3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin")) 
            ws['L3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['L3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top= Side(border_style="thin"), bottom = Side(border_style="thin"))                        

            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 10

            ws.row_dimensions[1].height = 25



            cont+=1

        nombre_archivo = "ReporteAutorExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")

        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


# -------------------- REUTILIZAR CENSO ------------------------

def votingIdSet():
    lista=[]
    for census in Census.objects.all():
        lista.append(census.voting_id)
    conjunto=set(lista)
    return conjunto

def reuseCensus(request):
    query = request.GET['q']
    c = request.GET['census']
    print(c)
    query = int(query)

    Census.objects.update(voting_id=query)
    census = Census.objects.filter(voting_id__icontains=query).order_by('-voter_id')
    return render(request, "census_reused.html", {'census':census, 'page_name':'Reuse Results'})

"""
class reuseContext(ListView):
    model = Census
    template_name = 'filterCensus.html'
    context_object_name = 'census'
    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs)"""